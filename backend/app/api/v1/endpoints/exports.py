import csv
import io
from datetime import datetime
from html import escape
from pathlib import Path

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel, Field
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer

router = APIRouter()
LOGO_PATH = Path(__file__).resolve().parents[5] / "app" / "assets" / "images" / "nutricore-brand.png"


class ExportPayload(BaseModel):
    app_mode: str = Field(default="personal")
    patient_name: str = Field(default="Pengguna NutriCore AI")
    patient_group: str = Field(default="-")
    institution_name: str = Field(default="-")
    institution_address: str = Field(default="-")
    patient_address: str = Field(default="-")
    medical_record_number: str = Field(default="-")
    visit_number: str = Field(default="-")
    payment_type: str = Field(default="-")
    guarantor_name: str = Field(default="-")
    national_id: str = Field(default="-")
    bpjs_number: str = Field(default="-")
    referral_source: str = Field(default="-")
    birth_date: str = Field(default="-")
    visit_date: str = Field(default="-")
    visit_time: str = Field(default="-")
    printed_at: str = Field(default="-")
    age: str = Field(default="-")
    weight_kg: str = Field(default="-")
    height_cm: str = Field(default="-")
    activity_level: str = Field(default="-")
    goal: str = Field(default="-")
    medical_conditions: str = Field(default="-")
    summary: str = Field(default="-")
    recommendations: list[str] = Field(default_factory=list)
    notes: str = Field(default="Dokumen ini dapat digunakan sebagai catatan tracking pribadi maupun bahan diskusi untuk petugas gizi dan tenaga medis.")
    report_type: str = Field(default="personal")
    clinician_name: str = Field(default="-")
    nutritionist_name: str = Field(default="-")
    follow_up_plan: str = Field(default="-")
    calorie_target_kcal: str = Field(default="-")
    protein_g: str = Field(default="-")
    carbs_g: str = Field(default="-")
    fat_g: str = Field(default="-")
    sodium_mg: str = Field(default="-")
    fluid_ml: str = Field(default="-")
    soap_subjective: str = Field(default="-")
    soap_objective: str = Field(default="-")
    soap_assessment: str = Field(default="-")
    soap_plan: str = Field(default="-")
    icd10_summary: str = Field(default="-")
    report_profile_title: str = Field(default="-")
    report_profile_focus: str = Field(default="-")


REPORT_TITLES = {
    "personal": "Catatan Tracking Pribadi",
    "nutritionist": "Laporan Asuhan Gizi Ringkas",
    "medical": "Ringkasan Klinis Pendamping",
}


def _report_title(report_type: str) -> str:
    return REPORT_TITLES.get(report_type, REPORT_TITLES["personal"])


def _build_filename(ext: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return f"nutricore-summary-{timestamp}.{ext}"


def _is_institution_mode(payload: ExportPayload) -> bool:
    return payload.app_mode == "institution"


def _pdf_bytes(payload: ExportPayload) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title="NutriCore AI Summary Export",
        author="dr Theresia AYH",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name="TitleCustom",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=22,
        leading=28,
        textColor=colors.HexColor("#102A6D"),
        spaceAfter=8,
    )
    body_style = ParagraphStyle(
        name="BodyCustom",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=11,
        leading=16,
        textColor=colors.HexColor("#10233F"),
        spaceAfter=6,
    )
    heading_style = ParagraphStyle(
        name="HeadingCustom",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=18,
        textColor=colors.HexColor("#1F4E94"),
        spaceBefore=8,
        spaceAfter=8,
    )

    story = []
    if LOGO_PATH.exists():
        story.append(Image(str(LOGO_PATH), width=52 * mm, height=20 * mm))
        story.append(Spacer(1, 4))
    if _is_institution_mode(payload):
        story.extend(
            [
                Paragraph(escape(payload.institution_name), title_style),
                Paragraph(escape(payload.institution_address), body_style),
                Paragraph("Dokumen Nutrisi & Klinik NutriCore AI", heading_style),
            ]
        )
    else:
        story.append(Paragraph("NutriCore AI", title_style))

    story.extend(
        [
            Paragraph(_report_title(payload.report_type), heading_style),
            Paragraph(f"<b>Tanggal kunjungan:</b> {escape(payload.visit_date)}", body_style),
            Paragraph(f"<b>Waktu kunjungan:</b> {escape(payload.visit_time)}", body_style),
            Paragraph(f"<b>Dicetak / diunduh:</b> {escape(payload.printed_at)}", body_style),
            Paragraph(f"<b>Nama:</b> {escape(payload.patient_name)}", body_style),
            Paragraph(f"<b>Kelompok pasien:</b> {escape(payload.patient_group)}", body_style),
            Paragraph(f"<b>Tanggal lahir:</b> {escape(payload.birth_date)}", body_style),
            Paragraph(f"<b>Usia:</b> {escape(payload.age)}", body_style),
            Paragraph(f"<b>Alamat pasien:</b> {escape(payload.patient_address)}", body_style),
        ]
    )
    if _is_institution_mode(payload):
        story.extend(
            [
                Paragraph(f"<b>No. rekam medis:</b> {escape(payload.medical_record_number)}", body_style),
                Paragraph(f"<b>No. kunjungan:</b> {escape(payload.visit_number)}", body_style),
                Paragraph(f"<b>Jenis pembayaran / penjamin:</b> {escape(payload.payment_type)}", body_style),
                Paragraph(f"<b>Nama penjamin:</b> {escape(payload.guarantor_name)}", body_style),
                Paragraph(f"<b>NIK:</b> {escape(payload.national_id)}", body_style),
                Paragraph(f"<b>No. BPJS:</b> {escape(payload.bpjs_number)}", body_style),
                Paragraph(f"<b>Rujukan / konsulan dari:</b> {escape(payload.referral_source)}", body_style),
            ]
        )
    story.extend(
        [
            Paragraph(f"<b>Berat badan:</b> {escape(payload.weight_kg)} kg", body_style),
            Paragraph(f"<b>Tinggi badan:</b> {escape(payload.height_cm)} cm", body_style),
            Paragraph(f"<b>Aktivitas:</b> {escape(payload.activity_level)}", body_style),
            Paragraph(f"<b>Tujuan:</b> {escape(payload.goal)}", body_style),
            Paragraph(f"<b>Kondisi medis:</b> {escape(payload.medical_conditions)}", body_style),
            Paragraph(f"<b>Jenis laporan:</b> {escape(_report_title(payload.report_type))}", body_style),
            Paragraph(f"<b>Target kalori:</b> {escape(payload.calorie_target_kcal)} kcal", body_style),
            Paragraph(
                f"<b>Makro:</b> Protein {escape(payload.protein_g)} g | Karbohidrat {escape(payload.carbs_g)} g | Lemak {escape(payload.fat_g)} g",
                body_style,
            ),
            Paragraph(
                f"<b>Natrium / Cairan:</b> {escape(payload.sodium_mg)} mg | {escape(payload.fluid_ml)} mL",
                body_style,
            ),
            Spacer(1, 8),
            Paragraph("Ringkasan", heading_style),
            Paragraph(escape(payload.summary), body_style),
            Spacer(1, 8),
            Paragraph("Rekomendasi", heading_style),
        ]
    )

    if payload.recommendations:
        for item in payload.recommendations:
            story.append(Paragraph(f"• {escape(item)}", body_style))
    else:
        story.append(Paragraph("Belum ada rekomendasi tambahan.", body_style))

    story.extend(
        [
            Spacer(1, 8),
            Paragraph("SOAP", heading_style),
            Paragraph(f"<b>S:</b> {escape(payload.soap_subjective)}", body_style),
            Paragraph(f"<b>O:</b> {escape(payload.soap_objective)}", body_style),
            Paragraph(f"<b>A:</b> {escape(payload.soap_assessment)}", body_style),
            Paragraph(f"<b>P:</b> {escape(payload.soap_plan)}", body_style),
            Paragraph("ICD-10 Pendamping", heading_style),
            Paragraph(escape(payload.icd10_summary), body_style),
            Paragraph("Profil Laporan Klinis", heading_style),
            Paragraph(f"<b>Template:</b> {escape(payload.report_profile_title)}", body_style),
            Paragraph(f"<b>Fokus:</b> {escape(payload.report_profile_focus)}", body_style),
            Spacer(1, 8),
            Paragraph("Catatan", heading_style),
            Paragraph(escape(payload.notes), body_style),
            Paragraph("Rencana Tindak Lanjut", heading_style),
            Paragraph(escape(payload.follow_up_plan), body_style),
            Paragraph("Tanda Tangan", heading_style),
            Paragraph(f"Petugas Gizi: ____________________  ({escape(payload.nutritionist_name)})", body_style),
            Paragraph(f"Petugas Medis: ____________________  ({escape(payload.clinician_name)})", body_style),
            Spacer(1, 8),
            Paragraph("Developer: dr Theresia AYH | April 2026", body_style),
        ]
    )

    doc.build(story)
    return buffer.getvalue()


def _csv_bytes(payload: ExportPayload) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Field", "Value"])
    writer.writerow(["Jenis laporan", _report_title(payload.report_type)])
    writer.writerow(["Mode aplikasi", payload.app_mode])
    writer.writerow(["Instansi", payload.institution_name])
    writer.writerow(["Alamat instansi", payload.institution_address])
    writer.writerow(["Alamat pasien", payload.patient_address])
    writer.writerow(["No. rekam medis", payload.medical_record_number])
    writer.writerow(["No. kunjungan", payload.visit_number])
    writer.writerow(["Jenis pembayaran / penjamin", payload.payment_type])
    writer.writerow(["Nama penjamin", payload.guarantor_name])
    writer.writerow(["NIK", payload.national_id])
    writer.writerow(["No. BPJS", payload.bpjs_number])
    writer.writerow(["Rujukan / konsulan", payload.referral_source])
    writer.writerow(["Tanggal lahir", payload.birth_date])
    writer.writerow(["Tanggal kunjungan", payload.visit_date])
    writer.writerow(["Waktu kunjungan", payload.visit_time])
    writer.writerow(["Dicetak / diunduh", payload.printed_at])
    writer.writerow(["Kelompok pasien", payload.patient_group])
    writer.writerow(["Nama", payload.patient_name])
    writer.writerow(["Usia", payload.age])
    writer.writerow(["Berat badan (kg)", payload.weight_kg])
    writer.writerow(["Tinggi badan (cm)", payload.height_cm])
    writer.writerow(["Aktivitas", payload.activity_level])
    writer.writerow(["Tujuan", payload.goal])
    writer.writerow(["Kondisi medis", payload.medical_conditions])
    writer.writerow(["Target kalori (kcal)", payload.calorie_target_kcal])
    writer.writerow(["Protein (g)", payload.protein_g])
    writer.writerow(["Karbohidrat (g)", payload.carbs_g])
    writer.writerow(["Lemak (g)", payload.fat_g])
    writer.writerow(["Natrium (mg)", payload.sodium_mg])
    writer.writerow(["Cairan (mL)", payload.fluid_ml])
    writer.writerow(["Ringkasan", payload.summary])
    writer.writerow(["SOAP - S", payload.soap_subjective])
    writer.writerow(["SOAP - O", payload.soap_objective])
    writer.writerow(["SOAP - A", payload.soap_assessment])
    writer.writerow(["SOAP - P", payload.soap_plan])
    writer.writerow(["ICD-10", payload.icd10_summary])
    writer.writerow(["Template klinis", payload.report_profile_title])
    writer.writerow(["Fokus laporan", payload.report_profile_focus])
    for index, item in enumerate(payload.recommendations, start=1):
        writer.writerow([f"Rekomendasi {index}", item])
    writer.writerow(["Catatan", payload.notes])
    writer.writerow(["Rencana tindak lanjut", payload.follow_up_plan])
    writer.writerow(["Petugas gizi", payload.nutritionist_name])
    writer.writerow(["Petugas medis", payload.clinician_name])
    return buffer.getvalue().encode("utf-8")


def _doc_bytes(payload: ExportPayload) -> bytes:
    recommendations = "".join(f"<li>{escape(item)}</li>" for item in payload.recommendations) or "<li>Belum ada rekomendasi tambahan.</li>"
    html = f"""
    <html>
      <head>
        <meta charset="utf-8" />
        <title>NutriCore AI Summary</title>
        <style>
          body {{
            font-family: Arial, sans-serif;
            color: #10233F;
            margin: 32px;
          }}
          h1, h2 {{
            color: #102A6D;
          }}
          table {{
            border-collapse: collapse;
            width: 100%;
            margin-bottom: 20px;
          }}
          td, th {{
            border: 1px solid #DCE5EE;
            padding: 8px;
            text-align: left;
          }}
          th {{
            background: #EEF4F8;
          }}
        </style>
      </head>
      <body>
        <img src="{LOGO_PATH.as_uri() if LOGO_PATH.exists() else ''}" alt="NutriCore AI" style="height:70px;margin-bottom:12px;" />
        <h1>NutriCore AI</h1>
        <h2>{escape(_report_title(payload.report_type))}</h2>
        <table>
          <tr><th>Field</th><th>Value</th></tr>
          <tr><td>Mode aplikasi</td><td>{escape(payload.app_mode)}</td></tr>
          <tr><td>Instansi</td><td>{escape(payload.institution_name)}</td></tr>
          <tr><td>Alamat instansi</td><td>{escape(payload.institution_address)}</td></tr>
          <tr><td>Alamat pasien</td><td>{escape(payload.patient_address)}</td></tr>
          <tr><td>No. rekam medis</td><td>{escape(payload.medical_record_number)}</td></tr>
          <tr><td>No. kunjungan</td><td>{escape(payload.visit_number)}</td></tr>
          <tr><td>Jenis pembayaran / penjamin</td><td>{escape(payload.payment_type)}</td></tr>
          <tr><td>Nama penjamin</td><td>{escape(payload.guarantor_name)}</td></tr>
          <tr><td>NIK</td><td>{escape(payload.national_id)}</td></tr>
          <tr><td>No. BPJS</td><td>{escape(payload.bpjs_number)}</td></tr>
          <tr><td>Rujukan / konsulan</td><td>{escape(payload.referral_source)}</td></tr>
          <tr><td>Tanggal lahir</td><td>{escape(payload.birth_date)}</td></tr>
          <tr><td>Tanggal kunjungan</td><td>{escape(payload.visit_date)}</td></tr>
          <tr><td>Waktu kunjungan</td><td>{escape(payload.visit_time)}</td></tr>
          <tr><td>Dicetak / diunduh</td><td>{escape(payload.printed_at)}</td></tr>
          <tr><td>Jenis laporan</td><td>{escape(_report_title(payload.report_type))}</td></tr>
          <tr><td>Nama</td><td>{escape(payload.patient_name)}</td></tr>
          <tr><td>Kelompok pasien</td><td>{escape(payload.patient_group)}</td></tr>
          <tr><td>Usia</td><td>{escape(payload.age)}</td></tr>
          <tr><td>Berat badan</td><td>{escape(payload.weight_kg)} kg</td></tr>
          <tr><td>Tinggi badan</td><td>{escape(payload.height_cm)} cm</td></tr>
          <tr><td>Aktivitas</td><td>{escape(payload.activity_level)}</td></tr>
          <tr><td>Tujuan</td><td>{escape(payload.goal)}</td></tr>
          <tr><td>Kondisi medis</td><td>{escape(payload.medical_conditions)}</td></tr>
          <tr><td>Target kalori</td><td>{escape(payload.calorie_target_kcal)} kcal</td></tr>
          <tr><td>Protein</td><td>{escape(payload.protein_g)} g</td></tr>
          <tr><td>Karbohidrat</td><td>{escape(payload.carbs_g)} g</td></tr>
          <tr><td>Lemak</td><td>{escape(payload.fat_g)} g</td></tr>
          <tr><td>Natrium</td><td>{escape(payload.sodium_mg)} mg</td></tr>
          <tr><td>Cairan</td><td>{escape(payload.fluid_ml)} mL</td></tr>
        </table>
        <h2>Ringkasan</h2>
        <p>{escape(payload.summary)}</p>
        <h2>Rekomendasi</h2>
        <ul>{recommendations}</ul>
        <h2>SOAP</h2>
        <p><strong>S:</strong> {escape(payload.soap_subjective)}</p>
        <p><strong>O:</strong> {escape(payload.soap_objective)}</p>
        <p><strong>A:</strong> {escape(payload.soap_assessment)}</p>
        <p><strong>P:</strong> {escape(payload.soap_plan)}</p>
        <h2>ICD-10 Pendamping</h2>
        <p>{escape(payload.icd10_summary)}</p>
        <h2>Profil Laporan Klinis</h2>
        <p><strong>Template:</strong> {escape(payload.report_profile_title)}</p>
        <p><strong>Fokus:</strong> {escape(payload.report_profile_focus)}</p>
        <h2>Catatan</h2>
        <p>{escape(payload.notes)}</p>
        <h2>Rencana Tindak Lanjut</h2>
        <p>{escape(payload.follow_up_plan)}</p>
        <h2>Tanda Tangan</h2>
        <p><strong>Petugas Gizi:</strong> ____________________ ({escape(payload.nutritionist_name)})</p>
        <p><strong>Petugas Medis:</strong> ____________________ ({escape(payload.clinician_name)})</p>
        <p><strong>Developer:</strong> dr Theresia AYH | April 2026</p>
      </body>
    </html>
    """
    return html.encode("utf-8")


def _xlsx_bytes(payload: ExportPayload) -> bytes:
    workbook = Workbook()
    thin = Side(style="thin", color="DCE5EE")
    header_fill = PatternFill("solid", fgColor="EEF4F8")
    section_fill = PatternFill("solid", fgColor="EAF6EC")
    bold_font = Font(bold=True, color="102A6D")

    def style_header(row):
        for cell in row:
            cell.font = bold_font
            cell.fill = header_fill
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    def style_cells(ws):
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    overview = workbook.active
    overview.title = "Overview"
    overview.append(["Field", "Value"])
    style_header(overview[1])
    overview_rows = [
        ["Jenis laporan", _report_title(payload.report_type)],
        ["Mode aplikasi", payload.app_mode],
        ["Instansi", payload.institution_name],
        ["Alamat instansi", payload.institution_address],
        ["Alamat pasien", payload.patient_address],
        ["No. rekam medis", payload.medical_record_number],
        ["No. kunjungan", payload.visit_number],
        ["Jenis pembayaran / penjamin", payload.payment_type],
        ["Nama penjamin", payload.guarantor_name],
        ["NIK", payload.national_id],
        ["No. BPJS", payload.bpjs_number],
        ["Rujukan / konsulan", payload.referral_source],
        ["Tanggal lahir", payload.birth_date],
        ["Tanggal kunjungan", payload.visit_date],
        ["Waktu kunjungan", payload.visit_time],
        ["Dicetak / diunduh", payload.printed_at],
        ["Kelompok pasien", payload.patient_group],
        ["Nama", payload.patient_name],
        ["Usia", payload.age],
        ["Berat badan (kg)", payload.weight_kg],
        ["Tinggi badan (cm)", payload.height_cm],
        ["Aktivitas", payload.activity_level],
        ["Tujuan", payload.goal],
        ["Kondisi medis", payload.medical_conditions],
        ["Ringkasan", payload.summary],
        ["SOAP - S", payload.soap_subjective],
        ["SOAP - O", payload.soap_objective],
        ["SOAP - A", payload.soap_assessment],
        ["SOAP - P", payload.soap_plan],
        ["ICD-10", payload.icd10_summary],
        ["Template klinis", payload.report_profile_title],
        ["Fokus laporan", payload.report_profile_focus],
        ["Catatan", payload.notes],
        ["Rencana tindak lanjut", payload.follow_up_plan],
        ["Petugas gizi", payload.nutritionist_name],
        ["Petugas medis", payload.clinician_name],
    ]
    for row in overview_rows:
        overview.append(row)
    overview.column_dimensions["A"].width = 28
    overview.column_dimensions["B"].width = 70
    style_cells(overview)

    nutrition = workbook.create_sheet("Nutrition Targets")
    nutrition.append(["Parameter", "Nilai"])
    style_header(nutrition[1])
    for row in [
        ["Kalori target (kcal)", payload.calorie_target_kcal],
        ["Protein (g)", payload.protein_g],
        ["Karbohidrat (g)", payload.carbs_g],
        ["Lemak (g)", payload.fat_g],
        ["Natrium (mg)", payload.sodium_mg],
        ["Cairan (mL)", payload.fluid_ml],
    ]:
        nutrition.append(row)
    nutrition.column_dimensions["A"].width = 28
    nutrition.column_dimensions["B"].width = 28
    style_cells(nutrition)

    recommendations = workbook.create_sheet("Recommendations")
    recommendations.append(["No", "Rekomendasi"])
    style_header(recommendations[1])
    if payload.recommendations:
        for index, item in enumerate(payload.recommendations, start=1):
            recommendations.append([index, item])
    else:
        recommendations.append([1, "Belum ada rekomendasi tambahan."])
    recommendations.column_dimensions["A"].width = 8
    recommendations.column_dimensions["B"].width = 90
    style_cells(recommendations)

    clinical = workbook.create_sheet("Clinical Notes")
    clinical.append(["Bagian", "Isi"])
    style_header(clinical[1])
    for row in [
        ["Tujuan laporan", _report_title(payload.report_type)],
        ["Template klinis", payload.report_profile_title],
        ["Fokus laporan", payload.report_profile_focus],
        ["SOAP - S", payload.soap_subjective],
        ["SOAP - O", payload.soap_objective],
        ["SOAP - A", payload.soap_assessment],
        ["SOAP - P", payload.soap_plan],
        ["ICD-10", payload.icd10_summary],
        ["Catatan", payload.notes],
        ["Rencana tindak lanjut", payload.follow_up_plan],
        ["Petugas gizi", payload.nutritionist_name],
        ["Petugas medis", payload.clinician_name],
        ["Developer", "dr Theresia AYH | April 2026"],
    ]:
        clinical.append(row)
    clinical.column_dimensions["A"].width = 26
    clinical.column_dimensions["B"].width = 88
    style_cells(clinical)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


@router.post("/summary")
def export_summary(
    payload: ExportPayload,
    format: str = Query(..., pattern="^(pdf|csv|doc|xlsx)$"),
):
    if format == "pdf":
        content = _pdf_bytes(payload)
        media_type = "application/pdf"
        filename = _build_filename("pdf")
    elif format == "csv":
        content = _csv_bytes(payload)
        media_type = "text/csv; charset=utf-8"
        filename = _build_filename("csv")
    elif format == "doc":
        content = _doc_bytes(payload)
        media_type = "application/msword"
        filename = _build_filename("doc")
    elif format == "xlsx":
        content = _xlsx_bytes(payload)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = _build_filename("xlsx")
    else:
        raise HTTPException(status_code=400, detail="Unsupported export format")

    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
