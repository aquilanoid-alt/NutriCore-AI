from __future__ import annotations

import math
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Dict

from openpyxl import load_workbook


DATA_DIR = Path(__file__).resolve().parents[1] / "data" / "who"


@dataclass(frozen=True)
class WhoReferenceRow:
    key: float
    l_value: float
    m_value: float
    s_value: float
    extra_values: Dict[str, float]


@dataclass(frozen=True)
class WhoReferenceTable:
    filename: str
    key_label: str
    step: float
    rows: tuple[WhoReferenceRow, ...]

    def nearest_row(self, input_value: float) -> WhoReferenceRow | None:
        if not self.rows:
            return None
        row = min(self.rows, key=lambda item: abs(item.key - input_value))
        max_distance = max(self.step / 2, 0.5)
        if abs(row.key - input_value) > max_distance:
            return None
        return row


@lru_cache(maxsize=None)
def load_who_table(filename: str) -> WhoReferenceTable:
    workbook = load_workbook(DATA_DIR / filename, data_only=True, read_only=True)
    worksheet = workbook.active
    headers = [str(item).strip() if item is not None else "" for item in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]

    key_index = 0
    l_index = headers.index("L")
    m_index = headers.index("M")
    s_index = headers.index("S")

    rows: list[WhoReferenceRow] = []
    for raw_row in worksheet.iter_rows(min_row=2, values_only=True):
        if raw_row[key_index] is None:
            continue
        rows.append(
            WhoReferenceRow(
                key=float(raw_row[key_index]),
                l_value=float(raw_row[l_index]),
                m_value=float(raw_row[m_index]),
                s_value=float(raw_row[s_index]),
                extra_values={
                    headers[index]: float(value)
                    for index, value in enumerate(raw_row)
                    if headers[index] and value is not None and index not in {key_index, l_index, m_index, s_index}
                },
            )
        )

    step = 1.0
    if len(rows) > 1:
        step = round(abs(rows[1].key - rows[0].key), 4) or 1.0

    return WhoReferenceTable(
        filename=filename,
        key_label=headers[key_index] or "key",
        step=step,
        rows=tuple(rows),
    )


def compute_zscore(value: float, reference_row: WhoReferenceRow) -> float:
    if value <= 0:
        raise ValueError("Anthropometry value must be greater than zero.")
    if reference_row.l_value == 0:
        return round(math.log(value / reference_row.m_value) / reference_row.s_value, 2)
    raw = ((value / reference_row.m_value) ** reference_row.l_value - 1) / (reference_row.l_value * reference_row.s_value)
    return round(raw, 2)
